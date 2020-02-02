import pandas as pd

from openpyxl.utils.dataframe import dataframe_to_rows
from shutil import copyfile
from openpyxl import load_workbook

template_file = 'C:\\work\\newoperplan\\testfiles\\Шаблон дефицитов.xlsx' # Has a header in row 1 already
output_file = 'C:\\work\\newoperplan\\testfiles\\Result.xlsx' # What we are saving the template as

# Copy Template.xlsx as Result.xlsx
copyfile(template_file, output_file)

d = {'col1': [1, 2], 'col2': [3, 4]}
df = pd.DataFrame(data=d)

print(df)

# Load the workbook and access the sheet we'll paste into
# wb = load_workbook(output_file)
# ws = wb['Дефицит']
# ws['A8']
# for r in dataframe_to_rows(df, index=False, header=False):
#     ws.append(r)

# book = load_workbook(output_file)
# writer = pd.ExcelWriter(output_file, engine='openpyxl') 
# writer.book = book
# writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

# df.to_excel(writer, index=False, sheet_name='Дефицит', startrow=5, header=None)

# writer.save()

# writer = pd.ExcelWriter(output_file, engine='openpyxl')
# df.to_excel(writer, index=False, sheet_name='Дефицит', startrow=6, header=None)
# writer.save()

# wb.save(output_file)

##############################################################################
#
# An example of converting a Pandas dataframe to an xlsx file
# with column formats using Pandas and XlsxWriter.
#
# Copyright 2013-2020, John McNamara, jmcnamara@cpan.org
#

# Create a Pandas dataframe from some data.
df = pd.DataFrame({'Numbers':    [1010, 2020, 3030, 2020, 1515, 3030, 4545],
                   'Percentage': [.1,   .2,   .33,  .25,  .5,   .75,  .45 ],
})

# Create a Pandas Excel writer using XlsxWriter as the engine.
writer = pd.ExcelWriter(output_file, engine='xlsxwriter')

# Convert the dataframe to an XlsxWriter Excel object.
df.to_excel(writer, sheet_name='Sheet1')

# Get the xlsxwriter workbook and worksheet objects.
workbook  = writer.book
worksheet = writer.sheets['Sheet1']

# Add some cell formats.
format1 = workbook.add_format({'num_format': '#,##0.00'})
format2 = workbook.add_format({'num_format': '0%'})

# Note: It isn't possible to format any cells that already have a format such
# as the index or headers or any cells that contain dates or datetimes.

# Set the column width and format.
worksheet.set_column('B:B', 18, format1)

# Set the format but not the column width.
worksheet.set_column('C:C', None, format2)

# Close the Pandas Excel writer and output the Excel file.
writer.save()