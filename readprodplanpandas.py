from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict
# ID	Готово	Отгрузка	Продукция	Контрагент	№ Заказа	№ СЗ	Заказ	цех	Статус	нр	кр	ц	вып.		пн

days_of_week = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']

in_id_name = 'ID'
shop_name = 'цех'

col_names = ['ID', 'цех']
col_names_index = []

periods_list = ['Ц', 'М', 'К', 'З']
series_list = ['О', 'Z']


class ReadFindError(Exception):
   pass

def get_dates_list(ws, row_num, col_num):
    row_num = row_num-1
    dates_list = []
    for col in range(col_num, ws.max_column+1):
        date_str = ws.cell(row=row_num, column=col).value
        if isinstance(date_str, datetime):
            dates_list.append(date_str)
        else:
            if not dates_list:
                raise ReadFindError('Что-то не так со строкой дат, значение: "%s" (строка: %s, столбец:%s)' % (date_str, row_num, col))
            print('Что-то не так с датами, последняя %s, (строка: %s, столбец:%s)' % (dates_list[-1], row_num, col))
            return dates_list
    return dates_list

def find_day(ws):
    # находим столбец с первым днем
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']:
                return [cell.row, cell.column]
    raise ReadFindError('Не нашел столбец "%s"' % day_list)

def find_names_col_num(ws, row_num, col_num, find_text):
    # находим номера нужных столбцов
    for col in range(1, col_num):
        if find_text == ws.cell(row=row_num, column=col).value:
            return col
    raise ReadFindError('Не нашел столбец "%s"' % find_text)

def get_word_dates(row, dates_list, periods_list, first_col):
    shop_dates = defaultdict(list)
    
    for i in range(first_col, len(row)):
        val = row[i-1].value
        if val:
            shop_dates[val].append(dates_list[i-first_col])
    
    for key, value in shop_dates.items():
        if key in periods_list:
            shop_dates[key] = [value[0], value[-1]]
    
    return shop_dates

def to_dict(file_path):
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb['График']
    try:
        first_day_coords = find_day(ws)
        print(first_day_coords)
        # for col_name in col_names:
        #     col_names_index.append(find_names_col_num(ws, first_day_coords[0], first_day_coords[1], col_name))
        # dates_list = get_dates_list(ws, first_day_coords[0], first_day_coords[1])
    except ReadFindError as ind:
        print('Ошибка: %s' % ind)
        exit(0)

    # import time
    # def timing():
    #     # Счетчик времени, таймер
    #     start_time = time.time()
    #     return lambda x: print("[{:>7.2f}с.] {}".format(time.time() - start_time, x))
    
    # t = timing() 

    # full_dates = []
    # # for row in range(first_day_coords[0]+1, ws.max_row+1):

    #     # full_dates.append(get_word_dates(ws, periods_list, series_list, dates_list, col_names_index[0], col_names_index[1], row, first_day_coords[1], ws.max_column+1))
    
    # for row in ws.iter_rows(min_row=first_day_coords[0]+1, max_col=ws.max_column+1, max_row=ws.max_row+1):
    # # for row in ws.iter_rows(min_row=115, max_col=ws.max_column+1, max_row=118):
    #     zz = get_word_dates(row, dates_list, periods_list, first_day_coords[1])
    #     if zz:
    #         full_dates.append(zz)
    # t("{:>5} Сделал".format('oo'))


    # # for i in full_dates:
    # #     for ss, kk in i.items():
    # #         print(ss, kk)

    # # тут запускаем функцию поиска нужных букв

def main():
    to_dict('testfiles\\График план производства\\План производства.xlsx')

if __name__ == "__main__":
    main()