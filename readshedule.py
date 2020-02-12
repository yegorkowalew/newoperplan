
"""
Графики ПДО

"""
import os
from openpyxl import load_workbook
import platform
from datetime import datetime
import pandas as pd

def creation_date(path_to_file):
    if platform.system() == 'Windows':
        creation_date = datetime.utcfromtimestamp(os.path.getctime(path_to_file))
        modification_date = datetime.utcfromtimestamp(os.path.getmtime(path_to_file))
        return creation_date, modification_date
    else:
        stat = os.stat(path_to_file)
        try:
            return stat.st_birthtime
        except AttributeError:
            creation_date = datetime.utcfromtimestamp(stat.st_mtime)
            return creation_date, None

def shedule_find_file(in_folder):
    extentions = ['xlsx']
    needfiles = []
    for root, _, nfiles in os.walk(in_folder):
        for filee in nfiles:
            if filee.split('.')[-1] in extentions:
                needfiles.append(os.path.join(root, filee))
    return needfiles

def get_dispatcher_from_path(file_path):
    return file_path.partition('Диспетчер')[2].partition(' - ')[2].partition('\\')[0]

def shedule_files_read(files):
    """
    Парсим имена графиков
    """
    error_files = []
    correct_files = []
    for xls_file in files:
        try:
            wb = load_workbook(filename=xls_file, read_only=True)
            ws = wb['График']
            cell_value = ws.cell(row=3, column=3).value
            if (len(str(cell_value)) != 7) and (len(str(cell_value)) != 9):
                # print(cell_value)
                error_files.append({
                    # 'dispatcher':xls_file.split('\\')[-3].split(' - ')[-1],
                    'dispatcher':get_dispatcher_from_path(xls_file),
                    'error':'Order number: %s in file does not match valid number' % (cell_value),
                    'file_path':xls_file
                })
            else:
                c_m_dates = creation_date(xls_file)
                correct_files.append({
                    # 'dispatcher':xls_file.split('\\')[-2].split(' - ')[-1],
                    'dispatcher':get_dispatcher_from_path(xls_file),
                    'file_path':xls_file,
                    'order_no':cell_value,
                    'date_creation': c_m_dates[0],
                    'date_modification': c_m_dates[1],
                    })
        except BaseException as identifier:
            # print(identifier)
            # print(xls_file.split('\\')[-3].split(' - ')[-1])
            error_files.append({
                # 'dispatcher':xls_file.split('\\')[-3].split(' - ')[-1],
                'dispatcher':get_dispatcher_from_path(xls_file),
                'file_path':xls_file,
                'error':identifier,
            })
    return correct_files, error_files

def sheduleWorker(folder):
    correct_files, error_files = shedule_files_read(shedule_find_file(folder))
    correct_files = pd.DataFrame.from_dict(correct_files)
    error_files = pd.DataFrame.from_dict(error_files)
    return correct_files, error_files

if __name__ == "__main__":
    from settings import SHEDULE_FOLDER
    correct_files, error_files = sheduleWorker(SHEDULE_FOLDER)
    correct_files.to_excel('testfiles\\Correct_Shedule_Files.xlsx')
    error_files.to_excel('testfiles\\Failure_Shedule_Files.xlsx')