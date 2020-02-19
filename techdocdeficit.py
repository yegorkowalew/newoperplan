import pandas as pd
import os
from settings import TODAY, TECH_DOC_BASE_FILE, TECH_DOC_DEFICIT_FOLDER, TECH_DOC_DAILY_REPORT_FOLDER
from openpyxl import load_workbook

def create_dataframe(dflist):
    df = pd.concat(dflist, axis=1, sort=False)
    df = df.loc[(df['ready_status'] == False) & (df['produced'] == True)]
        # №№СЗ
        # № Заказа
        # Контрагент
        # Продукция
        # Отгрузка
        # КВ Дата начала отсчета
        # КВ Дата выдачи по плану
        # КВ Дата выдачи по факту
        # КВ Разница дней
        # КВ Комментарий
        # ОС Дата начала отсчета
        # ОС Дата выдачи по плану
        # ОС Дата выдачи по факту
        # ОС Разница дней
        # ОС Комментарий
        # КД Дата начала отсчета
        # КД Дата выдачи по плану
        # КД Дата выдачи по факту
        # КД Разница дней
        # КД Комментарий
        
    df = df[[
        'sn_no', # №№СЗ
        'order_no', # № Заказа
        'counterparty', # Контрагент
        'product_name', # Продукция
        # 'shipment_from',
        'shipment_before', # Отгрузка
        # 'amount',
        'sn_date', # Дата начала комплектовочных, отгрузочных, чертежей,
        
        # КВ Дата начала отсчета 'sn_date'
        'pickup_plan_date_f',# КВ Дата выдачи по плану
        'pickup_date', # КВ Дата выдачи по факту
        # 'dispatcher_pickup_date', # Диспетчер отметевший дату выдачи по факту, пока не нужен
        'pickup_issue', # Нужна документация или нет
        'dispatcher_pickup_issue', # Диспетчер отметивший нужна документация или нет
        # КВ Разница дней
        # КВ Комментарий

        # ОС Дата начала отсчета 'sn_date'
        'shipping_plan_date_f', # ОС Дата выдачи по плану
        'shipping_date', # ОС Дата выдачи по факту
        'dispatcher_shipping_date', # Диспетчер отметевший дату выдачи по факту, пока не нужен
        'shipping_issue', # Нужна документация или нет
        'dispatcher_shipping_issue', # Диспетчер отметивший нужна документация или нет
        # ОС Разница дней
        # ОС Комментарий

        'design_plan_date_f',# КД Дата выдачи по плану
    ]].copy()
    df = df.dropna(subset=['sn_no'])
    return df

def worker_tech_doc(df, deficite_df):
    # Создаю столбцы с датами начала отсчета
    df['pickup_sn_date'] = df['sn_date']
    df['shipping_sn_date'] = df['sn_date']
    df['design_sn_date'] = df['sn_date']
    
    # Столбцы с разницей дней
    df['pickup_days'] = None
    df['shipping_days'] = None
    df['design_days'] = None

    # Столбцы с комментарием
    df['pickup_comment'] = None
    df['shipping_comment'] = None
    df['design_comment'] = None
    df['pickup_issue'] = df['pickup_issue'].fillna(True).astype(bool)

    # df['design_date'] = None

    print(deficite_df)
    df = pd.merge(df, deficite_df, on='order_no', how='outer')

    # Если не нужны, ставим в факт дату сз, в коммент пишем "Не нужны"
    def get_counterparty(sn_date, plan_date_f, date, days, comment, issue):
        # 'sn_date' # Начало отсчета
        # 'plan_date_f',# КВ Дата выдачи по плану
        # 'date', # КВ Дата выдачи по факту
        # 'days', # Дней
        # 'comment' # Комментарий
        # 'issue' # Нужна документация или нет
        if issue == False:
            comment = 'Не нужны'
            date = plan_date_f
        
        days = (plan_date_f - date).days

        if days > 0:
            comment = 'Раньше на %sдн.' % days

        if days < 0:
            comment = 'Позже на %sдн.' % abs(days)

        if days == 0:
            comment = 'В день по плану'

        if pd.isnull(date):
            days = (plan_date_f - TODAY).days
            if days > 0:
                comment = 'До выдачи %sдн.' % days
                days = '! %s' % (plan_date_f - TODAY).days
            if days < 0:
                comment = 'Просрочка %sдн.' % abs(days)
                days = '! %s' % abs((plan_date_f - TODAY).days)
            if days == 0:
                comment = 'Выдача сегодня'
                days = '! %s' % (plan_date_f - TODAY).days

        return sn_date, plan_date_f, date, days, comment

    def get_design_counterparty(sn_date, plan_date_f, date, days, comment, document_count, without_date_count):
        # 'sn_date' # Начало отсчета
        # 'plan_date_f',# КВ Дата выдачи по плану
        # 'date', # КВ Дата выдачи по факту
        # 'days', # Дней
        # 'comment' # Комментарий
        # 'issue' # Нужна документация или нет
        # if issue == False:
        #     comment = 'Не нужны'
        #     date = plan_date_f
        
        # days = (plan_date_f - date).days

        # if days > 0:
        #     comment = 'Раньше на %sдн.' % days

        # if days < 0:
        #     comment = 'Позже на %sдн.' % abs(days)

        # if days == 0:
        #     comment = 'В день по плану'

        # if pd.isnull(date):
        #     days = (plan_date_f - TODAY).days
        #     if days > 0:
        #         comment = 'До выдачи %sдн.' % days
        #         days = '! %s' % (plan_date_f - TODAY).days
        #     if days < 0:
        #         comment = 'Просрочка %sдн.' % abs(days)
        #         days = '! %s' % abs((plan_date_f - TODAY).days)
        #     if days == 0:
        #         comment = 'Выдача сегодня'
        #         days = '! %s' % (plan_date_f - TODAY).days
        if not pd.isnull(document_count):
            if not pd.isnull(date):
                days = (plan_date_f - date).days
            comment = 'Не выданы: %s из %s' % (int(without_date_count), int(document_count))
        return sn_date, plan_date_f, date, days, comment


    for index, row in df.iterrows():
        df.loc[index, 'pickup_sn_date'], df.loc[index, 'pickup_plan_date_f'], df.loc[index, 'pickup_date'], df.loc[index, 'pickup_days'], df.loc[index, 'pickup_comment'] = get_counterparty(row['pickup_sn_date'],row['pickup_plan_date_f'],row['pickup_date'],row['pickup_days'],row['pickup_comment'],row['pickup_issue'])
        df.loc[index, 'shipping_sn_date'], df.loc[index, 'shipping_plan_date_f'], df.loc[index, 'shipping_date'], df.loc[index, 'shipping_days'], df.loc[index, 'shipping_comment'] = get_counterparty(row['shipping_sn_date'],row['shipping_plan_date_f'],row['shipping_date'],row['shipping_days'],row['shipping_comment'],row['shipping_issue'])
        df.loc[index, 'design_sn_date'], \
        df.loc[index, 'design_plan_date_f'], \
        df.loc[index, 'design_date'], \
        df.loc[index, 'design_days'], \
        df.loc[index, 'design_comment'] = get_design_counterparty(
            row['design_sn_date'],
            row['design_plan_date_f'],
            row['design_date'],
            row['design_days'],
            row['design_comment'],
            # row['shipping_issue']
            row['document_count'],
            row['without_date_count'],

            )

    # Пересобираю df в правильном порядке столбцов
    df = df[[
        'sn_no', # №№СЗ
        'order_no', # № Заказа
        'counterparty', # Контрагент
        'product_name', # Продукция
        # 'shipment_from',
        # 'shipment_before', # Отгрузка
        # 'amount',
        # 'sn_date', # Дата начала комплектовочных, отгрузочных, чертежей,
        
        'pickup_sn_date', # КВ Дата начала отсчета 'sn_date'
        'pickup_plan_date_f',# КВ Дата выдачи по плану
        'pickup_date', # КВ Дата выдачи по факту
        'pickup_days', # Разница в днях
        'pickup_comment', # Комментарий
        # 'dispatcher_pickup_date', # Диспетчер отметевший дату выдачи по факту, пока не нужен
        # 'pickup_issue', # Нужна документация или нет
        # 'dispatcher_pickup_issue', # Диспетчер отметивший нужна документация или нет

        'shipping_sn_date', # ОС Дата начала отсчета 'sn_date'
        'shipping_plan_date_f', # ОС Дата выдачи по плану
        'shipping_date', # ОС Дата выдачи по факту
        'shipping_days', # Разница в днях
        'shipping_comment', # Комментарий
        # 'dispatcher_shipping_date', # Диспетчер отметевший дату выдачи по факту, пока не нужен
        # 'shipping_issue', # Нужна документация или нет
        # 'dispatcher_shipping_issue', # Диспетчер отметивший нужна документация или нет

        'design_sn_date',
        'design_plan_date_f', # КД Дата выдачи по плану
        'design_date', # ОС Дата выдачи по факту
        'design_days', # Разница в днях
        'design_comment', # Комментарий
    ]]

    return df

def techdocFindFile(in_folder):
    tree = os.walk(in_folder)
    files_list = []
    for address, dirs, files in tree:
        for fl in files:
            if not '~$' in fl.split('.')[0] and fl.split('.')[-1] == 'xlsx':
                files_list.append(os.path.join(address, fl))
    return files_list

def reportReadFile(path_file):
    try:
        df = pd.read_excel(
            path_file,
            sheet_name="Дефицит",
            header=3,
            usecols = [
                'Наименование',
                'Дата выдачи',
            ],

            parse_dates = [
                'Дата выдачи',
            ],
        )
    except Exception as ind:
        # logger.error("inDocumentReadFile error with file: %s - %s" % (file_name, ind))
        print("inDocumentReadFile error with file: %s - %s" % (path_file, ind))
    else:
        return df

def get_dispatcher_from_path(file_path):
    # Есть высокая вероятность того что с именем оператора можем не угадать
    return file_path.split('\\')[-2]

def techdocbase(in_folder):
    files_list = techdocFindFile(in_folder)
    df_arr = []
    for report in files_list:
        dispatcher_design_date = get_dispatcher_from_path(report)
        wb = load_workbook(filename=report, read_only=True)
        ws = wb['Дефицит']
        order_no = ws.cell(row=2, column=2).value
        df = reportReadFile(report)
        df['order_no'] = order_no
        df['dispatcher_design_date'] = dispatcher_design_date
        df_arr.append(df)
    df = pd.concat(df_arr, ignore_index=True, sort=False)

    base_file = df.rename({'Наименование': 'Наименование', 'Дата выдачи': 'Дата выдачи', 'order_no': 'Заказ №', 'dispatcher_design_date': 'Диспетчер'}, axis='columns')
    base_file = base_file[[
        'Диспетчер',
        'Заказ №',
        'Наименование',
        'Дата выдачи'
    ]]

    base_file.to_excel(TECH_DOC_BASE_FILE, index=False)
    df = df.rename({'Наименование': 'detail', 'Дата выдачи': 'design_date', 'order_no': 'order_no', 'dispatcher_design_date': 'dispatcher_design_date'}, axis='columns')
    df['order_no'] = df['order_no'].astype(str)
    return df

def techdocdeficite(df):
    df = df[df['design_date'].isnull()] 
    df = df.rename({'detail': 'Наименование', 'design_date': 'Дата выдачи', 'order_no': 'Заказ №', 'dispatcher_design_date': 'Диспетчер'}, axis='columns')
    df = df[[
        'Диспетчер',
        'Заказ №',
        'Наименование',
    ]]
    filename = '%s - дефицит КД.xlsx' % TODAY.strftime('%Y.%m.%d %H-%M')
    df.to_excel(os.path.join(TECH_DOC_DAILY_REPORT_FOLDER, filename), index=False)

def techdocreport(df):
    import numpy as np
    df =df.drop('detail', axis=1)
    df['dispatcher_design_date'] = df['dispatcher_design_date'] + '||'
    df['document_count'] = 1
    df['without_date_count'] = df['design_date'].apply(lambda x: 1 if pd.isnull(x) else 0 )
    dates = df.groupby('order_no').agg({'design_date': [np.max], 'dispatcher_design_date': [np.sum], 'document_count': [np.sum], 'without_date_count': [np.sum]})
    dates = dates.reset_index(level='order_no', col_level=1, col_fill='order_no')
    dates.columns = [col[0] for col in dates.columns]
    dates['dispatcher_design_date'] = dates['dispatcher_design_date'].apply(lambda x: x.split('||')[0])
    df['order_no'] = df['order_no'].astype(str)
    return dates


if __name__ == "__main__":
    from settings import READY_FILE, SN_FILE, IN_DOCUMENT_FILE, IN_DOCUMENT_FOLDER, PRODUCTION_PLAN_FILE, SHEDULE_FOLDER
    from readready import readyReadFile
    from readservicenote import serviceNoteReadFile
    from readproductionplan import productionPlanReadFile
    from readindocument import worker
    from appenddata import appendDataWorker

    import time
    def timing():
        # Счетчик времени, таймер
        start_time = time.time()
        return lambda x: print("[{:>7.2f}с.] {}".format(time.time() - start_time, x))
    
    t = timing() #[  76.51с.]       Конец выполнения
    
    from multiprocessing import Pool
    from multiprocessing.dummy import Pool as ThreadPool
    pool = ThreadPool()

    results = []
    result_objects = [
        pool.apply_async(readyReadFile, (READY_FILE,)),
        pool.apply_async(serviceNoteReadFile, (SN_FILE,)),
        pool.apply_async(productionPlanReadFile, (PRODUCTION_PLAN_FILE,)),
        pool.apply_async(worker, (IN_DOCUMENT_FILE, IN_DOCUMENT_FOLDER,)),
        ]
    
    all_df = create_dataframe([result_objects[0].get(), result_objects[1].get(), result_objects[2].get(), result_objects[3].get()])
    df = techdocbase(TECH_DOC_DEFICIT_FOLDER)
    deficite = techdocreport(df)

    all_df = worker_tech_doc(all_df, deficite)
    all_df.to_excel('testfiles\\DeficitTechnicalDocumentation.xlsx')

    # techdocdeficite(df)
    
    # df_merge_col = pd.merge(all_df, deficite, on='order_no', how='outer')
    
    # df_merge_col.to_excel('testfiles\\test3.xlsx')
    t("{:>5} Конец выполнения".format(''))
